"""Convert the master Excel into an intermediate JSON without mutating the workbook.

Preserves original column values. Splits the existing combined column
`유형 / 난이도` on the last ` / ` separator into type and difficulty tokens.
Does not invent Gold labels for fields that are absent from Excel.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_EXCEL = Path(r"c:\Users\rudck\Documents\카카오톡 받은 파일\연금_AI_Agent_100개_마스터테스트.xlsx")
SHEET_CASES = "전체 100개"
SHEET_SETS = "세트 요약"

# Columns as printed in Excel row 1 — do not rename meanings.
COL_TEST = "Test"
COL_SET = "세트"
COL_TYPE_DIFF = "유형 / 난이도"
COL_QUESTION = "테스트 질문"
COL_EXPECTED = "기대 답변"


def split_type_difficulty(combined: str) -> tuple[str, str]:
    """Parse Excel's combined `유형 / 난이도` cell. Difficulty is the final token."""
    text = str(combined or "").strip()
    if " / " in text:
        left, right = text.rsplit(" / ", 1)
        return left.strip(), right.strip()
    return text, ""


def derive_family(type_label: str) -> str:
    """Coarse family for reporting only, derived from Excel 유형 text."""
    t = type_label or ""
    if any(k in t for k in ("복합 Routing", "복합", "교차", "대조")) and "Routing" in t:
        return "Compound"
    if any(k in t for k in ("정보 부족", "모호")):
        return "Abstention"
    if any(k in t for k in ("잘못된 전제", "전제 오류", "교정")):
        return "Correction"
    if any(k in t for k in ("세액공제", "세제 연산", "분리과세", "과세이연", "Engine", "경계값", "연산", "계산")):
        return "Calculation"
    if any(k in t for k in ("상품", "SQL", "총보수", "TDF", "모펀드", "편입", "합성")):
        return "Product"
    if any(k in t for k in ("추천",)):
        return "Recommendation"
    if any(k in t for k in ("법률", "세법", "규약", "자격", "중도인출", "담보", "지연이자")):
        return "Legal"
    if any(k in t for k in ("디폴트", "옵트인", "실물이전", "통지", "만기", "대기")):
        return "Procedure"
    if any(k in t for k in ("VPC", "인프라", "전산")):
        return "Infrastructure"
    if any(k in t for k in ("RAG", "제도", "실무", "교육", "재정", "DB ")):
        return "Document"
    return "Other"


def adapter_eval_hints(type_label: str, expected_answer: str) -> dict[str, Any]:
    """Evaluation policy hints derived from Excel 유형 + 기대 답변 wording.

    These are NOT Excel Gold label columns. They are adapter-side scoring policy
    so metrics (routing/source/abstention) can be computed without inventing
    missing spreadsheet fields.
    """
    t = type_label or ""
    e = expected_answer or ""
    hints: dict[str, Any] = {
        "source": "adapter_policy_from_excel_type_and_expected_answer",
        "answer_required": True,
        "require_clarify": False,
        "require_abstention_or_clarify": False,
        "require_correction": False,
        "require_postgres": False,
        "require_enterprise_document": False,
        "require_calculation_route": False,
        "require_law_or_document": False,
        "safe_stop_forbidden": True,
        "no_arbitrary_product": False,
        "route_families": [],
        "source_families": [],
    }

    if any(k in t for k in ("정보 부족", "모호")) or any(
        k in e for k in ("추가 정보", "단정 추천하면 안", "부족하므로")
    ):
        hints["require_abstention_or_clarify"] = True
        hints["require_clarify"] = True
        hints["safe_stop_forbidden"] = False
        hints["no_arbitrary_product"] = True
        hints["route_families"] = ["product", "conversation"]
        hints["source_families"] = ["product", "postgres", "none"]

    if any(k in t for k in ("잘못된 전제", "전제 오류", "교정")) or e.strip().startswith("아님"):
        hints["require_correction"] = True
        hints["route_families"] = ["calculation", "document", "document+law", "law"]
        hints["source_families"] = ["calculation", "enterprise_document", "law"]

    if any(k in t for k in ("상품", "SQL", "총보수", "합성", "모펀드", "편입비율", "운용제한")):
        hints["require_postgres"] = True
        hints["route_families"] = list(dict.fromkeys([*(hints["route_families"] or []), "product"]))
        hints["source_families"] = list(
            dict.fromkeys([*(hints["source_families"] or []), "product", "postgres", "enterprise_document"])
        )
        if "RAG" in t or "모펀드" in t or "전략" in e:
            hints["require_enterprise_document"] = True

    if any(k in t for k in ("세액공제", "세제 연산", "분리과세", "과세이연", "Engine", "경계값", "초과납입")):
        hints["require_calculation_route"] = True
        hints["route_families"] = ["calculation", "calculation+law", "document+law"]
        hints["source_families"] = ["calculation", "law", "enterprise_document"]

    if any(k in t for k in ("제도", "실무", "절차", "디폴트", "옵트인", "실물이전", "교육", "재정", "규약", "법률", "세법")):
        if not hints["require_postgres"] and not hints["require_calculation_route"]:
            hints["require_enterprise_document"] = True
            hints["require_law_or_document"] = True
            hints["route_families"] = ["document", "document+law", "law"]
            hints["source_families"] = ["enterprise_rag", "enterprise_document", "law"]

    if "복합 Routing" in t:
        hints["route_families"] = ["document", "document+law", "product", "calculation", "law"]
        hints["source_families"] = [
            "enterprise_rag",
            "enterprise_document",
            "product",
            "postgres",
            "law",
            "calculation",
        ]

    if not hints["route_families"]:
        hints["route_families"] = ["document", "document+law", "law", "product", "calculation"]
    if not hints["source_families"]:
        hints["source_families"] = ["enterprise_document", "enterprise_rag", "law", "product", "postgres"]

    return hints


def extract_expected_numbers(expected_answer: str) -> list[str]:
    """Numeric tokens from Excel 기대 답변 for coverage checks (not invented)."""
    text = expected_answer or ""
    found = re.findall(r"\d{1,3}(?:,\d{3})+(?:\.\d+)?|\d+(?:\.\d+)?", text)
    cleaned: list[str] = []
    for item in found:
        digits = item.replace(",", "")
        # Keep decimals, thousands, codes (>=2 digits), skip lone list-like 1/2/3.
        if "." in item or "," in item or len(digits) >= 2:
            cleaned.append(item)
    return list(dict.fromkeys(cleaned))


def load_excel(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    if SHEET_CASES not in wb.sheetnames:
        raise ValueError(f"missing sheet {SHEET_CASES!r}; found {wb.sheetnames}")

    ws = wb[SHEET_CASES]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    expected_header = [COL_TEST, COL_SET, COL_TYPE_DIFF, COL_QUESTION, COL_EXPECTED]
    if header != expected_header:
        raise ValueError(f"unexpected header {header!r}; expected {expected_header!r}")

    cases: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, 6)]
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        test_id, set_name, type_diff, question, expected = values
        type_label, difficulty = split_type_difficulty(str(type_diff or ""))
        record = {
            "test_id": str(test_id).strip(),
            "세트": str(set_name or "").strip(),
            "유형_난이도_원문": str(type_diff or "").strip(),
            "유형": type_label,
            "난이도": difficulty,
            "테스트_질문": str(question or "").strip(),
            "기대_답변": str(expected or "").strip(),
            "family": derive_family(type_label),
            "expected_numbers_from_기대답변": extract_expected_numbers(str(expected or "")),
            "adapter_eval_hints": adapter_eval_hints(type_label, str(expected or "")),
        }
        cases.append(record)

    set_summary: list[dict[str, Any]] = []
    if SHEET_SETS in wb.sheetnames:
        s = wb[SHEET_SETS]
        set_header = [s.cell(1, c).value for c in range(1, s.max_column + 1)]
        for r in range(2, s.max_row + 1):
            row = {str(set_header[i]): s.cell(r, i + 1).value for i in range(len(set_header))}
            set_summary.append(row)

    return {
        "source_excel": str(path),
        "sheet": SHEET_CASES,
        "header": expected_header,
        "case_count": len(cases),
        "cases": cases,
        "세트_요약": set_summary,
        "notes": [
            "Original Excel was not modified.",
            "유형/난이도 split uses the existing combined column only.",
            "adapter_eval_hints are scoring policy, not Excel Gold label columns.",
            "No missing Gold fields were fabricated.",
        ],
    }


def write_intermediate(payload: dict[str, Any], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path


def validate_payload(payload: dict[str, Any]) -> list[str]:
    issues: list[str] = []
    cases = payload.get("cases") or []
    if len(cases) != 100:
        issues.append(f"expected 100 cases, got {len(cases)}")
    seen: set[str] = set()
    for case in cases:
        tid = case.get("test_id")
        if not tid:
            issues.append("missing test_id")
        elif tid in seen:
            issues.append(f"duplicate test_id:{tid}")
        else:
            seen.add(tid)
        if not case.get("테스트_질문"):
            issues.append(f"{tid}: empty 테스트_질문")
        if not case.get("기대_답변"):
            issues.append(f"{tid}: empty 기대_답변")
        if not case.get("난이도"):
            issues.append(f"{tid}: empty 난이도 after split of 유형 / 난이도")
    return issues


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--out", type=Path, default=Path("reports/gold100/gold100_cases.json"))
    args = parser.parse_args()
    payload = load_excel(args.excel)
    write_intermediate(payload, args.out)
    issues = validate_payload(payload)
    print(json.dumps({"cases": payload["case_count"], "issues": issues, "out": str(args.out)}, ensure_ascii=False))
    if issues:
        raise SystemExit(1)
